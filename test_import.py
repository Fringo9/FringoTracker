import openpyxl
from openpyxl import Workbook
from datetime import datetime

# Crea un nuovo workbook
wb = Workbook()
ws = wb.active
ws.title = "Finanze"

# Aggiungi intestazione con i mesi
ws['A1'] = "Voci"
ws['B1'] = "gen-24"
ws['C1'] = "feb-24"
ws['D1'] = "mar-24"

# Aggiungi alcuni dati di esempio
data = [
    ["Liquidi", 1000, 1100, 950],
    ["Banca", 5000, 5200, 5100],
    ["Credito", 300, 350, 400],
    ["Investimenti", 10000, 10500, 11000],
    ["Debito", -2000, -2100, -2000],
    ["Finanziamento", -5000, -5000, -5000],
    ["Totale", 9300, 10050, 10450],
]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Salva il file
wb.save("test_import.xlsx")
print("File di test creato: test_import.xlsx")
